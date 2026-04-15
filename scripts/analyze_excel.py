import openpyxl
import json

wb = openpyxl.load_workbook(r'\\172.16.2.100\Arquivo\Direcao Producao\Report diário produção.xlsm', read_only=True, data_only=True, keep_links=False)
result = {}
for s in wb.sheetnames:
    ws = wb[s]
    rows = list(ws.iter_rows(max_row=min(ws.max_row or 0, 50), values_only=True))
    result[s] = {
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'first_rows': []
    }
    for r in rows[:5]:
        result[s]['first_rows'].append([str(c)[:60] if c is not None else '' for c in r])
wb.close()

for name, info in result.items():
    print(f"\n{'='*60}")
    print(f"SHEET: {name}  |  rows: {info['max_row']}  |  cols: {info['max_col']}")
    print(f"{'='*60}")
    for i, row in enumerate(info['first_rows']):
        print(f"  Row {i+1}: {row}")
