import openpyxl

wb = openpyxl.load_workbook(
    r'C:/Users/Goncalo.Barata/Desktop/App stocks/reference/Report diário produção.xlsm',
    read_only=True, data_only=True, keep_links=False)

ws = wb['Tabuas_desperdicios_aux']
print(f"Tabuas_desperdicios_aux: {ws.max_row} rows x {ws.max_column} cols")
print("=" * 120)

# Dump full contents
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    cells = [str(c) if c is not None else '' for c in row]
    non_empty = [(j + 1, c) for j, c in enumerate(cells) if c != '']
    if non_empty:
        print(f"R{i}: {non_empty}")

wb.close()
