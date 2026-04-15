import openpyxl

wb = openpyxl.load_workbook(r'\\172.16.2.100\Arquivo\Direcao Producao\Report diário produção.xlsm', read_only=True, data_only=False, keep_links=False)

# Main sheet formulas (rows 1-47)
ws = wb['Main']
print("="*80)
print("MAIN - FORMULAS (rows 1-47)")
print("="*80)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=47, max_col=20), 1):
    for j, cell in enumerate(row):
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  [{i},{j}] {cell.value[:120]}")

# Rolaria_aux (try lowercase)
for name in wb.sheetnames:
    if 'rolaria' in name.lower():
        ws = wb[name]
        print(f"\n{'='*80}")
        print(f"{name} - ALL FORMULAS")
        print(f"{'='*80}")
        for i, row in enumerate(ws.iter_rows(values_only=False), 1):
            for j, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  [{i},{j}] {cell.value[:150]}")

# MCF data - sample formulas from first 20 rows
ws = wb['MCF data']
print(f"\n{'='*80}")
print("MCF data - SAMPLE FORMULAS (rows 1-20)")
print("="*80)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=63, values_only=False), 1):
    for j, cell in enumerate(row):
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  [{i},{j}] {cell.value[:150]}")

# MainDataTable - sample formulas
ws = wb['MainDataTable']
print(f"\n{'='*80}")
print("MainDataTable - SAMPLE FORMULAS (rows 1-15)")
print("="*80)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, max_col=20, values_only=False), 1):
    for j, cell in enumerate(row):
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  [{i},{j}] {cell.value[:150]}")

wb.close()
