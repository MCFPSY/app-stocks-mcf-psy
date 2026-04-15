import openpyxl

wb = openpyxl.load_workbook(r'\\172.16.2.100\Arquivo\Direcao Producao\Report diário produção.xlsm', read_only=True, data_only=True, keep_links=False)

target_sheets = ['Main', 'MCF data', 'PSY data', 'MainDataTable', 'Rolaria_Aux', 'Tabuas_desperdicios_aux', 'Aux_benchmark']

for s in wb.sheetnames:
    ws = wb[s]
    print(f"SHEET: {s}  |  rows: {ws.max_row}  |  cols: {ws.max_column}")

print("\n" + "="*80)
print("DETAIL: Main (rows 1-47)")
print("="*80)
ws = wb['Main']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=47, values_only=True), 1):
    cells = [str(c)[:40] if c is not None else '' for c in row[:20]]
    non_empty = [(j,c) for j,c in enumerate(cells) if c]
    if non_empty:
        print(f"  R{i}: {non_empty}")

for sname in target_sheets:
    if sname == 'Main':
        continue
    if sname not in wb.sheetnames:
        print(f"\n*** SHEET '{sname}' NOT FOUND ***")
        continue
    ws = wb[sname]
    print(f"\n{'='*80}")
    print(f"DETAIL: {sname}  |  rows: {ws.max_row}  |  cols: {ws.max_column}")
    print(f"{'='*80}")
    # Show first 10 rows, limited cols
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=min(ws.max_column or 20, 20), values_only=True), 1):
        cells = [str(c)[:35] if c is not None else '' for c in row]
        non_empty = [(j,c) for j,c in enumerate(cells) if c]
        if non_empty:
            print(f"  R{i}: {non_empty}")

wb.close()
