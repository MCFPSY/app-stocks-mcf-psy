import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Goncalo.Barata\Desktop\App stocks\reference\Report diário produção.xlsm', read_only=True, data_only=True, keep_links=False)

# Aux_benchmark - full dump
ws = wb['Aux_benchmark']
print(f"=== Aux_benchmark: {ws.max_row} rows x {ws.max_column} cols ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    non_empty = [(j, c) for j, c in enumerate(row) if c is not None]
    if non_empty and i <= 30:
        print(f"R{i}: {non_empty}")
print("... (showing first 30 rows)")

# Also check Aux and Parametros sheets — might have production lines for MCF
for sname in ['Aux', 'Parametros', 'MCF_Produtividade']:
    if sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n=== {sname}: {ws.max_row}x{ws.max_column} ===")
        for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
            non_empty = [(j, c) for j, c in enumerate(row) if c is not None]
            if non_empty:
                print(f"R{i}: {non_empty[:10]}")

# Look at MCF data columns for line references
ws = wb['MCF data']
print(f"\n=== MCF data headers (rows 12-17, first 30 cols) ===")
for i, row in enumerate(ws.iter_rows(min_row=12, max_row=17, max_col=30, values_only=True), 12):
    non_empty = [(j, c) for j, c in enumerate(row) if c is not None]
    if non_empty:
        print(f"R{i}: {non_empty}")

wb.close()
